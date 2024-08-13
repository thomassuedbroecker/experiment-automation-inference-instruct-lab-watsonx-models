from modules.requests_watsonx import watsonx_prompt
from modules.requests_instructlab import instruct_prompt
import argparse
import openpyxl
from openpyxl.styles import Alignment

# ******************************************
# Functions
def get_questions(excel_input_file):            
            wb = openpyxl.load_workbook(excel_input_file)
            ws = wb.active
            rows = []
            for rdx, row in enumerate(ws.iter_rows(values_only=True)):
                if rdx:
                        rows.append(list(row))
                else:
                        header = row      
            return header, rows

def create_output_workbook (output_file):
        workbook = openpyxl.Workbook()
        worksheet = workbook.create_sheet("experiment_data")

        if 'Sheet1' in workbook.sheetnames:
                 workbook.remove( workbook['Sheet1'])
        if 'Sheet' in  workbook.sheetnames:
                 workbook.remove( workbook['Sheet'])
        
        # Data
        worksheet.title = "experiment_data"
        worksheet['A1'] = 'input_example'
        worksheet['B1'] = 'prompt'
        worksheet['C1'] = 'model'
        worksheet['D1'] = 'result'
        worksheet['E1'] = 'result_goodness'
        worksheet['F1'] = 'result_commentary'
        worksheet['G1'] = 'extra_model_version'
        worksheet['H1'] = 'extra_generated_token_count'
        worksheet['I1'] = 'extra_input_token_count'
        worksheet['J1'] = 'extra_stop_reason'

        # Save the workbook as a new Excel file
        workbook.save(output_file)
        return workbook

def write_to_output_workbook(outputfile, results):

        workbook = openpyxl.load_workbook(outputfile)
        
        j = 1
        for row in results:
                worksheet = workbook['experiment_data']

                worksheet.cell(row=(j+1), column=1).value = row[0] #'input_example'
                worksheet.cell(row=(j+1), column=2).value = row[1] #'prompt'
                worksheet.cell(row=(j+1), column=3).value = row[3] #'model'
                worksheet.cell(row=(j+1), column=4).value = row[2] #'result'
                worksheet.cell(row=(j+1), column=5).value = "Insert your goodness between '1 and 10'" #'result_goodness'
                worksheet.cell(row=(j+1), column=6).value = "Insert your comment" #'result_commentary'
                worksheet.cell(row=(j+1), column=7).value = row[4] #'extra_model_version'
                worksheet.cell(row=(j+1), column=8).value = row[5] #'extra_generated_token_count'
                worksheet.cell(row=(j+1), column=9).value = row[6] #'extra_input_token_count'
                worksheet.cell(row=(j+1), column=10).value = row[7] #'extra_stop_reason'
                j = j + 1
        
        worksheet = workbook['experiment_data']
        for row in worksheet.iter_rows():  
                for cell in row:      
                         cell.alignment = Alignment(wrapText=True,vertical='top')
        
        workbook.save(outputfile)
        return True

def run_wx_experiment(questions):
        results = []
        #print(f"***Questions:\n{questions}\n")
        for question in questions:
                #print(f"***Question:{question[0]}")
                response, validation = watsonx_prompt(question[0])
                if validation['status'] == True:
                        print(f"***Result:\n{response}\n")
                        item = [response['result']['question'],
                                response['result']['prompt'],
                                response['result']['generated_text'],
                                response['result']['model_id'],
                                response['result']['model_version'],
                                response['result']['generated_token_count'],
                                response['result']['input_token_count'],
                                response['result']['stop_reason']]
                        results.append(item)
                else:
                        print(f"ERROR: {validation}\n{response}\n")
                        item = [response['prompt'],
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR"]
                        results.append(item)
        return results

def run_instlab_experiment(questions):
        results = []
        #print(f"***Questions:\n{questions}\n")
        for question in questions:
                #print(f"***Question:{question[0]}")
                response, validation = instruct_prompt(question[0])
                if validation['status'] == True:
                        print(f"***Result:\n{response}\n")
                        item = [response['result']['question'],
                                response['result']['prompt'],
                                response['result']['generated_text'],
                                response['result']['model_id'],
                                response['result']['model_version'],
                                response['result']['generated_token_count'],
                                response['result']['input_token_count'],
                                response['result']['stop_reason']]
                        results.append(item)
                else:
                        print(f"ERROR: {validation}\n{response}\n")
                        item = [response['prompt'],
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR",
                                "ERROR"]
                        results.append(item)
        return results

# ******************************************
# Execution
def main(args):
        inputfile = args.inputfile
        outputfile = args.outputfile
        inference = args.inference

        header, questions = get_questions(inputfile)
        print(f"***Log: Questions:\n{questions}\n")

        if (inference == "watsonx"):
                if (len(header)==1):
                        results = run_wx_experiment(questions)
                        print(f"Results: {results}")
                        create_output_workbook(outputfile)
                        write_to_output_workbook(outputfile, results)
                else:
                        print(f"Please verify content in inputfile: {inputfile}")
        
        elif (inference == "instructlab"):
                if (len(header)==1):
                        results = run_instlab_experiment(questions)
                        print(f"Results: {results}")
                        create_output_workbook(outputfile)
                        write_to_output_workbook(outputfile, results)
                else:
                        print(f"Please verify content in inputfile: {inputfile}")
        else:
                print("Please select a valid inference: 'watsonx' or 'instructlab'?")              

if __name__ == "__main__":
    
    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--inputfile')
    parser.add_argument('-o', '--outputfile')
    parser.add_argument('-r', '--inference')
    args = parser.parse_args()

    print(f"Input:{args.inputfile}\nOutput:{args.outputfile}")
    
    # check that questions_file matches db_type
    if args.inputfile is None:
        print("Input file is missing!")
        exit()

    if args.outputfile is None:
        print("Output file is missing!")
        exit()

    if args.inference is None:
        print("Inference is missing: 'watsonx' or 'instructlab'?")
        exit()

    main(args)